import json
import yaml
from openpyxl import load_workbook


def _get_security_policy_yaml_data(value):
    _yaml_data_security_policy = {
        f"payload{payload_number}": {} for payload_number, _ in enumerate(value)
    }
    for index, data in enumerate(value):
        security_policy_entry = {"name": data["cws_policy_name"]}
        _yaml_data_security_policy["payload{}".format(index)].update(
            security_policy_entry
        )
    return _yaml_data_security_policy


def _get_url_filtering_policy_yaml_data(value):
    _yaml_data_url_filtering = {
        f"payload{payload_number}": {} for payload_number, _ in enumerate(value)
    }
    for index, data in enumerate(value):
        url_filtering_entry = {
            "ruleIndex": 0,
            "rule": {
                "name": data["url_filtering_rule_name"],
                "action": {"value": data["url_filtering_rule_action_value"]},
                "attributes": {
                    "fqdns": [
                        fqdn.strip()
                        for fqdn in data["url_filtering_rule_attributes_fqdns"].split(
                            ","
                        )
                    ],
                    "userEntity": {
                        "all": bool(
                            data["url_filtering_rule_attributes_user_Entity_all"]
                        ),
                        "userGroups": [
                            group.strip()
                            for group in data[
                                "url_filtering_rule_attributes_userEntity_userGroups"
                            ].split(",")
                        ],
                        "users": [
                            user.strip()
                            for user in data[
                                "url_filtering_rule_attributes_userEntity_users"
                            ].split(",")
                        ],
                    },
                },
            },
        }
        _yaml_data_url_filtering["payload{}".format(index)].update(url_filtering_entry)
    return _yaml_data_url_filtering


def _get_content_filtering_policy_yaml_data(value):
    _yaml_data_content_filtering = {
        f"payload{payload_number}": {} for payload_number, _ in enumerate(value)
    }
    for index, data in enumerate(value):
        content_filtering_entry = {
            "ruleIndex": 0,
            "rule": {
                "name": data["content_filtering_rule_name"],
                "action": {
                    "allowedSites": {
                        "print": data["rule_action_allowedSites_print"],
                        "action": data["rule_action_allowedSites_action"],
                        "skipAV": data["rule_action_allowedSites_skipAV"],
                        "viewer": data["rule_action_allowedSites_viewer"],
                        "passwordAction": data[
                            "rule_action_allowedSites_passwordAction"
                        ],
                    }
                },
                "attributes": {
                    "fqdns": data["rule_attributes_fqdns"],
                    "fileType": data["rule_attributes_fileType"],
                    "fileValues": data["rule_attributes_fileValues"],
                    "categories": data["rule_attributes_categories"],
                    "userEntity": {
                        "all": data["rule_attributes_userEntity_all"],
                        "userGroups": [data["rule_attributes_userEntity_userGroups_0"]],
                        "users": [data["rule_attributes_userEntity_users_0"]],
                    },
                    "transferType": data["rule_attributes_transferType"],
                },
            },
        }
        _yaml_data_content_filtering["payload{}".format(index)].update(
            content_filtering_entry
        )
    return _yaml_data_content_filtering


def _get_geo_filtering_policy_yaml_data(value):
    _yaml_data_geo_based_filtering = {
        f"payload{payload_number}": {} for payload_number, _ in enumerate(value)
    }
    for index, data in enumerate(value):
        geo_based_filtering_entry = {
            "ruleIndex": 0,
            "rule": {
                "name": data["rule_name"],
                "action": {
                    "action": data["rule_action_action"],
                    "captureLog": data["rule_action_captureLog"],
                },
                "attributes": {
                    "tags": [data["rule_attributes_tags_0"]],
                    "reason": data["rule_attributes_reason"],
                    "locations": [data["rule_attributes_locations_0"]],
                    "userEntity": {
                        "all": data["rule_attributes_userEntity_all"],
                        "users": [data["rule_attributes_userEntity_users"]],
                        "userGroups": [data["rule_attributes_userEntity_userGroups"]],
                    },
                },
            },
        }
        _yaml_data_geo_based_filtering["payload{}".format(index)].update(
            geo_based_filtering_entry
        )
    return _yaml_data_geo_based_filtering


def _get_content_inspection_policy_yaml_data(value):
    _yaml_data_content_inspection = {
        f"payload{payload_number}": {} for payload_number, _ in enumerate(value)
    }
    for index, data in enumerate(value):
        content_inspection_entry = {
            "ruleIndex": 0,
            "rule": {
                "name": data["rule_name"],
                "action": {"value": data["rule_action_value"]},
                "attributes": {
                    "modules": [
                        data["rule_attributes_modules_0"],
                        data["rule_attributes_modules_1"],
                    ],
                    "fileType": data["rule_attributes_fileType"],
                    "categories": [data["rule_attributes_categories_0"]],
                    "fileValues": [data["rule_attributes_fileValues_0"]],
                    "userEntity": {
                        "all": data["rule_attributes_userEntity_all"],
                        "users": [data["rule_attributes_userEntity_users"]],
                        "userGroups": [data["rule_attributes_userEntity_userGroups"]],
                    },
                    "contentType": data["rule_attributes_contentType"],
                    "transferType": [data["rule_attributes_transferType_0"]],
                },
            },
        }

        _yaml_data_content_inspection["payload{}".format(index)].update(
            content_inspection_entry
        )

    return _yaml_data_content_inspection


def _get_ssl_inspection_policy_yaml_data(value):
    _yaml_data_ssl_inspection = {
        f"payload{payload_number}": {} for payload_number, _ in enumerate(value)
    }
    for index, data in enumerate(value):
        ssl_inspection_entry = {
            "ruleIndex": 0,
            "rule": {
                "name": data["rule_name"],
                "action": {"value": data["rule_action_value"]},
                "attributes": {"categories": [data["rule_attributes_categories"]]},
            },
        }

        _yaml_data_ssl_inspection["payload{}".format(index)].update(
            ssl_inspection_entry
        )

    return _yaml_data_ssl_inspection


def _get_casb_policy_yaml_data(value):
    _yaml_data_casb = {
        f"payload{payload_number}": {} for payload_number, _ in enumerate(value)
    }
    for index, data in enumerate(value):
        casb_entry = {
            "ruleIndex": 0,
            "rule": {
                "name": data["rule_name"],
                "action": {
                    "browserAction": data["rule_browserAction"],
                    "commonControls": [
                        {"name": "Login", "allow": data["login_allow"]},
                        {"name": "Upload", "allow": data["upload_allow"]},
                        {"name": "Download", "allow": data["download_allow"]},
                    ],
                    "appSpecificControls": [
                        {"name": "Create", "allow": data["create_allow"]},
                        {"name": "Delete", "allow": data["delete_allow"]},
                        {"name": "Edit", "allow": data["edit_allow"]},
                        {"name": "Like", "allow": data["like_allow"]},
                        {"name": "Post", "allow": data["post_allow"]},
                        {"name": "Search", "allow": data["search_allow"]},
                        {"name": "Share", "allow": data["share_allow"]},
                    ],
                },
                "attributes": {
                    "appIds": [data["attributes_appIDs"]],
                    "userEntity": {
                        "all": data["rule_attributes_userEntity_all"],
                        "users": [data["rule_attributes_userEntity_users"]],
                        "userGroups": [data["rule_attributes_userEntity_userGroups"]],
                    },
                },
            },
        }

        _yaml_data_casb["payload{}".format(index)].update(casb_entry)

    return _yaml_data_casb


def _get_dlp_policy_yaml_data(value):
    _yaml_data_dlp = {
        f"payload{payload_number}": {} for payload_number, _ in enumerate(value)
    }
    for index, data in enumerate(value):
        dlp_entry = {
            "ruleIndex": 0,
            "rule": {
                "name": data["rule_name"],
                "action": {
                    "action": data["rule_action"],
                    "dlpDicts": ["fc97a143-fd85-11ec-a5dd-02d4f172e1b6"],
                    "docTypes": [{"name": "All", "fileType": "All"}],
                    "protocols": ["http", "https"],
                    "sendAlert": data["sendAlert"],
                    "inspectInputs": data["rule_name"],
                    "maxFileSizeUnit": "MB",
                    "sessionType": "all",
                    "dlpAuditors": [],
                    "maxFileSize": 50,
                    "inspectInputs": False,
                    "domains": [],
                    "dlpGroups": [],
                    "categories": []
                },
                "attributes": {
                    "userEntity": {"all": True, "users": [], "userGroups": []},
                    "notification": ""
                },
            },
        }

        _yaml_data_dlp["payload{}".format(index)].update(dlp_entry)

    return _yaml_data_dlp


def _convert_excel_to_yaml(excel_file):
    yaml_data = {
        "security_policy": {},
        "url_filtering_policy": {},
        "content_filtering_policy": {},
        "geo_filtering_policy": {},
        "content_inspection_policy": {},
        "ssl_inspection_policy": {},
        "casb_policy": {},
        "dlp_policy": {},
    }

    # Load the Excel file
    workbook = load_workbook(excel_file)

    # Create a dictionary to store the data for all sheets
    all_data = {}

    # Iterate over all sheet names
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Create a list to store the data for the current sheet
        sheet_data = []

        # Get the column headers (keys)
        headers = [cell.value for cell in sheet[1]]

        # Process each row of data starting from the second row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Assign the values to corresponding keys (column headers)
            row_data = {
                header: value if value is not None and value != "" else []
                for header, value in zip(headers, row)
            }

            # Append the row data to the sheet_data list
            sheet_data.append(row_data)

        # Add the sheet_data list to the all_data dictionary under the sheet name key
        all_data[sheet_name] = sheet_data

    for key, value in all_data.items():
        if key == "url_filtering_policy":
            yaml_data_url_filtering_policy = _get_url_filtering_policy_yaml_data(value)
            yaml_data["url_filtering_policy"].update(yaml_data_url_filtering_policy)

        elif key == "security_policy":
            yaml_data_security_policy = _get_security_policy_yaml_data(value)
            yaml_data["security_policy"].update(yaml_data_security_policy)

        elif key == "content_filtering_policy":
            yaml_data_content_filtering_policy = (
                _get_content_filtering_policy_yaml_data(value)
            )
            yaml_data["content_filtering_policy"].update(
                yaml_data_content_filtering_policy
            )

        elif key == "geo_filtering_policy":
            yaml_data_geo_filtering_policy = _get_geo_filtering_policy_yaml_data(value)
            yaml_data["geo_filtering_policy"].update(yaml_data_geo_filtering_policy)

        elif key == "content_inspection_policy":
            yaml_data_content_inspection_policy = (
                _get_content_inspection_policy_yaml_data(value)
            )
            yaml_data["content_inspection_policy"].update(
                yaml_data_content_inspection_policy
            )

        elif key == "ssl_inspection_policy":
            yaml_data_ssl_inspection_policy = _get_ssl_inspection_policy_yaml_data(
                value
            )
            yaml_data["ssl_inspection_policy"].update(yaml_data_ssl_inspection_policy)

        elif key == "casb_policy":
            yaml_data_casb_policy = _get_casb_policy_yaml_data(value)
            yaml_data["casb_policy"].update(yaml_data_casb_policy)

        elif key == "dlp_policy":
            yaml_data_dlp_policy = _get_dlp_policy_yaml_data(value)
            yaml_data["dlp_policy"].update(yaml_data_dlp_policy)

    return yaml_data
